from config_new import (
    q, a, EV_velocity, gamma, gamma_l, EV_cost, tol, plot_enabled,
    GV_cost, w_dv, w_ev, theta, battery_threshold, V, N, Q_EV
)
from gurobipy import Model, GRB, quicksum
import itertools
from itertools import permutations
from openpyxl import load_workbook
import pandas as pd
import os
import re
import numpy as np
from itertools import combinations
from collections import deque
import heapq
import random
import time
import matplotlib.pyplot as plt
from dataclasses import dataclass
np.random.seed(42)

def ev_travel_cost(route):
    q[0]=0
    b = 1
    l = 0
    for i in range(len(route)-1):
        l+=q[route[i]]
        b = b - (a[route[i],route[i+1]]/EV_velocity)*(gamma+gamma_l*l) 
        if b < battery_threshold:
            raise ValueError(f"Battery level too low (below threshold of {battery_threshold}): {b} in route {route}")
    cost = 260*EV_cost*(1-b)
    return cost

def gv_tsp_cost(route):
    cost_GV = a[(route[0],route[1])]*GV_cost
    l = 0
    q[0]=0
    try:
        route = eval(route)
    except:
        for i in range(1,len(route)-1):
            l+=q[route[i]]
            cost_GV += a[(route[i],route[i+1])]*GV_cost*l
        return cost_GV
    
def check_values(d, tolerance=1e-5):
    for key, value in d.items():
        if not (abs(value - 0) <= tolerance or abs(value - 1) <= tolerance):
            return False
    return True

def partial_path(label,current_node):
    node = current_node
    path = []
    while label:
        path.append(label.node)
        label = label.parent
    if node in path[1:]:
        return True
    else: return False

def reconstruct_path(label):
    path = []
    while label:
        path.append(label.node)
        label = label.parent
    path = path[::-1]
    if path[0]=='s':
        path[0] = 0
    if path[-1]=='t':
        path[-1]= 0
    return path

def construct_tour(edges):
    next_node_map = {i: j for i, j in edges}
    # Start from any node (choosing the first node from the first edge)
    start_node = edges[0][0]
    route = [start_node]
    
    current_node = start_node
    while True:
        next_node = next_node_map.get(current_node)
        if next_node is None or next_node == start_node:  # End of cycle
            break
        route.append(next_node)
        current_node = next_node

    route.append(start_node)  # Complete the cycle
    return route

def print_solution(final_model) -> None:
    model_vars = {var.VarName: var.X for var in final_model.getVars()}
    var_names = list(model_vars.keys())

    solution_routes = []
    payments = {}

    for var in var_names:
        value = round(model_vars[var], 2) # Get the variable's value

        if var.startswith("p_"):  # Payment variables
            payments[int(var.split("_")[-1])] = value  # Convert index to int
        elif value != 0 and var.startswith("y_r_["):  # Route variables
            print(f"{var}: {value}")
            solution_routes.append(list(eval(var[5:-1])))  # Extract tuple inside []

    # Print results
    c_r = {tuple(item):0 for item in solution_routes}
    for item in solution_routes:
        if len(item) > 3:
            c_r[tuple(item)] = ev_travel_cost(item)

    total_dv_miles_traveled = 0
    total_ev_miles_traveled = 0

    for route in solution_routes:
        if len(route) == 3:
            for i in range(0,len(route)-1):
                total_dv_miles_traveled += a[(route[i], route[i+1])]
        else:
            for i in range(0,len(route)-1):
                total_ev_miles_traveled += a[(route[i], route[i+1])]
    print(solution_routes)
    print(f"Total miles cost (ev+dv): {total_dv_miles_traveled*w_dv + total_ev_miles_traveled*w_ev}")
    print(f"Total EV miles traveled: {total_ev_miles_traveled}")
    print(f"Total DV miles traveled: {total_dv_miles_traveled}")
    print(f"Total mniles traveled: {total_dv_miles_traveled + total_ev_miles_traveled}")
    print(f"Objective value: {final_model.getObjective().getValue()}")
    print(f"Objective value (manual): {total_dv_miles_traveled*w_dv + total_ev_miles_traveled*w_ev+theta*(sum(c_r.values())-sum(payments.values()))}")
    print("Total payment received:", sum(payments.values()))
    print("Total coalition cost:", sum(c_r.values()))
    print("Total Subsidy:", sum(c_r.values())-sum(payments.values()))
    print(f"Payments: {payments}")

    return final_model.getObjective().getValue(), total_dv_miles_traveled*w_dv + total_ev_miles_traveled*w_ev \
        , total_ev_miles_traveled, sum(payments.values()), sum(c_r.values())-sum(payments.values()), payments, solution_routes

def print_metadata(Total_CG_iteration, Total_RG_iteration, num_nodes_explored,
                  Total_RG_time, Total_CG_time, Total_RG_DP_time, Total_CG_DP_time,
                  Total_LP_time, tsp_cache_time, obj, root_obj_val, Total_num_lp):
    """
    Print a summary of branch-and-price performance stats.
    """
    print(f"Total CG iterations: {Total_CG_iteration}")
    print(f"Total RG iterations: {Total_RG_iteration}")
    print(f"Total nodes explored: {num_nodes_explored}")
    print(f"Total RG time: {Total_RG_time:.2f}")
    print(f"Total CG time: {Total_CG_time:.2f}")
    print(f"Total RG DP time: {Total_RG_DP_time:.2f}")
    print(f"Total CG DP time: {Total_CG_DP_time:.2f}")
    print(f"Total LP time: {Total_LP_time:.2f}")
    print(f"tsp cache time: {tsp_cache_time}")
    print(f"LP gap(%): {((obj - root_obj_val) / obj) * 100}")
    print(f"root_obj_val: {root_obj_val}")
    print(f"Total number of LPs solved: {Total_num_lp}")

def code_status(use_column_heuristic, always_generate_rows):
    if use_column_heuristic==False and always_generate_rows==True:
        code=1
    elif use_column_heuristic==False and always_generate_rows==False:
        code=2
    elif use_column_heuristic==True and always_generate_rows==False:
        code=3
    return code

def validate_solution(payments, tsp_memo, N, solution_routes):
    """
    Check solution validity: payment violations, IR violations,
    and compute EV travel costs for solution routes.
    """
    payments[0] = 0

    # --- Payment violations ---
    payment_flag = False
    for item in tsp_memo:
        pay = sum(payments[i] for i in item)
        if pay > tsp_memo[item][1] + 0.1:
            payment_flag = True
            print("Payment violation:")
            print("  Route:", item)
            print("  Total payments:", pay)
            print("  TSP cost:", tsp_memo[item][1])
    if not payment_flag:
        print("No payment violations")

    # --- IR violations ---
    ir_flag = False
    for i in N:
        if payments[i] > gv_tsp_cost((0, i, 0)) + 0.01:
            ir_flag = True
            print(f"IR violation at node {i}: payment={payments[i]}, IR cost={gv_tsp_cost((0, i, 0))}")
    if not ir_flag:
        print("No IR violations")

    # --- EV travel cost check ---
    for route in solution_routes:
        ev_travel_cost(route)

def generate_tsp_cache(N, k):
    print(f"\nCreating initial TSP memo... \ntotal {N}C{k} combinations\n")
    tsp_memo = {}
    global_tsp_memo = {}
    customers = list(range(1, N + 1)) 
    all_combinations = []
      
    for i in range(1, k + 1):
        for comb in combinations(customers, i):
            all_combinations.append((0,) + comb + (0,))
    for item in all_combinations:
        if sum(q[i] for i in item) <= Q_EV:
            candidate_route, tsp_cost = tsp_tour(item)
            tsp_memo[item] = (tuple(candidate_route),tsp_cost)
            global_tsp_memo[item] = (tuple(candidate_route),tsp_cost)
        else:
            candidate_route, tsp_cost = tsp_tour(item)
            global_tsp_memo[item] = (tuple(candidate_route),tsp_cost)

    print(f"\nFinished creating initial TSP memo\n")
    time.sleep(1)
    return tsp_memo, global_tsp_memo

def update_plot(outer_iter, lp_gap, iterations, lp_gaps):
    """Updates the LP gap plot dynamically."""
    if not plot_enabled:
        return  # Do nothing if plotting is disabled

    # Append new data
    iterations.append(outer_iter)
    lp_gaps.append(lp_gap)

    # Static figure and axis setup (only first time)
    if not hasattr(update_plot, "fig"):
        plt.ion()  # Enable interactive mode
        update_plot.fig, update_plot.ax = plt.subplots()
        update_plot.line, = update_plot.ax.plot([], [], marker='o', linestyle='-')

    # Update the plot
    update_plot.line.set_xdata(iterations)
    update_plot.line.set_ydata(lp_gaps)
    update_plot.ax.relim()  # Recompute limits
    update_plot.ax.autoscale_view()  # Adjust view
    update_plot.ax.set_xlabel("Outer Iteration")
    update_plot.ax.set_ylabel("LP Gap (%)")
    update_plot.ax.set_title("LP Gap vs. Outer Iteration")
    plt.draw()
    plt.pause(0.01)  # Pause for smooth updating

def make_stack(search_mode="fifo"):
    """
    Create a stack structure based on the specified search mode.
    Returns
    -------
    stack : deque | list
    push  : function push(node)
    pop   : function -> node
    """

    if search_mode == "fifo":
        stack = deque()
        def push(x): stack.append(x)
        def pop():  return stack.popleft()

    elif search_mode == "heap":
        stack = []
        def push(x): heapq.heappush(stack, x)
        def pop():  return heapq.heappop(stack)

    elif search_mode == "mixed":
        stack = []
        def push(x): heapq.heappush(stack, x)
        def pop():
            # 80% best-first, 20% random (same behavior as before)
            if random.random() < 0.8:
                return heapq.heappop(stack)
            idx = random.randint(0, len(stack) - 1)
            node = stack[idx]
            stack[idx] = stack[-1]
            stack.pop()
            heapq.heapify(stack)
            return node
    elif search_mode == "lifo":
        stack = []
        def push(x): stack.append(x)
        def pop():  return stack.pop()
    else:
        raise ValueError(f"Unknown SEARCH_MODE: {search_mode}")
    return stack, push, pop



def generate_all_possible_routes(N):
    
    all_routes = []

    def generate_k_degree_coalition(N, k):
        # Generate all combinations of k nodes
        combinations = list(itertools.combinations(N, k))
        # Generate all possible routes starting and ending at depot 0
        degree_k_coalition = [tuple([0] + list(comb) + [0]) for comb in combinations]
        return degree_k_coalition
    
    for item in N:
        all_routes.extend(generate_k_degree_coalition(N, item))
    
    return all_routes
def tsp_tour(route):
    
    if len(route) == 3 and route[0]==0:
        return route, a[(0,route[1])]* GV_cost * q[route[1]] + a[(route[1],0)] * GV_cost

    intermediate_nodes = route[1:-1]
    all_routes = [[0] + list(p) + [0] for p in permutations(intermediate_nodes)]
    routes_list = [tuple(all_routes) for all_routes in all_routes]
    route_cost = {}
    for item in routes_list:
        route_cost[tuple(item)] = gv_tsp_cost(item)

    model = Model("TSP")
    x = model.addVars(routes_list, vtype=GRB.BINARY, name="x")

    model.addConstr(quicksum(x[i] for i in routes_list) == 1)

    model.setObjective(quicksum(route_cost[i] * x[i] for i in routes_list), GRB.MINIMIZE)

    model.Params.OutputFlag = 0
    model.Params.MIPGap = 0.00000001
    model.optimize()
    
    # Extract the solution
    if model.status == GRB.OPTIMAL:
        for i in routes_list:
            if x[i].X > 0.5:
                tour = i
                break
        return tour, model.getObjective().getValue()
    
def save_to_excel(file_name, sheet_name, data):
    try:
        # Check if the file already exists
        if os.path.exists(file_name):
            # Load the existing workbook
            book = load_workbook(file_name)
            # Check if the sheet already exists
            if sheet_name in book.sheetnames:
                # Read the existing data in the sheet
                existing_data = pd.read_excel(file_name, sheet_name=sheet_name)
                # Append the new data to the existing data
                updated_data = pd.concat([existing_data, data], ignore_index=True)
            else:
                # If the sheet doesn't exist, use the new data
                updated_data = data

            # Write the updated data to the sheet
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                updated_data.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # If the file doesn't exist, create a new one with the data
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Data saved to {file_name} in sheet '{sheet_name}'.")
    except Exception as e:
        print(f"Error: {e}")

def update_config(node_value):
    # Path to the config_new.py file
    config_file = "New_codes/config_new.py"

    # Step 1: Update the NODES value in the config_new.py file
    with open(config_file, "r") as f:
        lines = f.readlines()

    with open(config_file, "w") as f:
        for line in lines:
            if line.startswith("NODES ="):
                f.write(f"NODES = {node_value}\n")  # Update NODES value
            else:
                f.write(line)

    # Step 3: Re-execute the rest of the code in the config_new module
    # This will execute the entire code from config_new.py again
    os.system('python3 New_codes/config_new.py')  # Executes the script after updating NODES


def create_excel_for_log_file(log_file):

    # Regular expression pattern to extract timestamp, description, and time value
    log_pattern = r"^(.*?) - (.*?): (\d+\.\d+)"

    # Lists to store parsed data
    timestamps = []
    descriptions = []
    time_values = []

    # Read the log file and extract data
    with open(log_file, "r") as file:
        for line in file:
            match = re.search(log_pattern, line)
            if match:
                timestamps.append(match.group(1))      # Timestamp
                descriptions.append(match.group(2))    # Description
                time_values.append(float(match.group(3)))  # Time value

    # Create a DataFrame
    df = pd.DataFrame({
        "Timestamp": timestamps,
        "Description": descriptions,
        "Time Taken": time_values
    })

    # Save to an Excel file
    log_file = log_file.split(".")[0]
    excel_filename = f"{log_file}.xlsx"
    df.to_excel(excel_filename, index=False)

    print(f"Data successfully saved to {excel_filename}")

class prize_collecting_tsp:
    def __init__(self, p_result=None, dual_values_delta=None, dual_values_subsidy=None, dual_values_IR=None, dual_values_vehicle=None):
        self.p_result = p_result
        self.dual_values_delta = dual_values_delta
        self.dual_values_subsidy = dual_values_subsidy
        self.dual_values_IR = dual_values_IR
        self.dual_values_vehicle = dual_values_vehicle

    def pctsp(self):
        # Decision variables
        self.m = Model("PrizeCollectingTSP")
        self.x = self.m.addVars(V, V, vtype=GRB.BINARY, name="x")      # arc used
        self.y = self.m.addVars(V, vtype=GRB.BINARY, name="y")         # node visited
        self.f = self.m.addVars(V, V, vtype=GRB.CONTINUOUS, lb=0.0, ub=Q_EV, name="f")  # flow on arc
        # Flow balance: each visited node must have exactly one in/out arc
        for i in V:
            self.m.addConstr(quicksum(self.x[i, j] for j in V if j != i) == self.y[i])
            self.m.addConstr(quicksum(self.x[j, i] for j in V if j != i) == self.y[i])

        # Depot must be visited
        self.m.addConstr(self.y[0] == 1)

        # Truck starts empty at depot
        self.m.addConstr(quicksum(self.f[0, j] for j in V if j != 0) == 0, name="DepotStartEmpty")

        # Truck returns to depot carrying total pickups collected
        self.m.addConstr(quicksum(self.f[j, 0] for j in V if j != 0) ==
                    quicksum(q[i] * self.y[i] for i in N),
                    name="DepotReturnFull")

        # Flow capacity: if arc is not used, no flow
        for i in V:
            for j in V:
                if i != j:
                    self.m.addConstr(self.f[i, j] <= Q_EV * self.x[i, j])

        # Flow conservation for pickups
        for i in N:  # customers only
            self.m.addConstr(
                quicksum(self.f[i, j] for j in V if j != i)
                - quicksum(self.f[j, i] for j in V if j != i)
                == q[i] * self.y[i]
            )
        return self.m

    def cg_pctsp(self):
        self.m = self.pctsp()
        self.b = self.m.addVars(V + ['t'], vtype=GRB.CONTINUOUS, ub = 1, lb = 0, name="b")         # battery level
        self.m.addConstr(self.b[0] == 1, name="DepotBatteryFull")                          # depot starts with full battery
        self.m.addConstrs(self.b[i] >= battery_threshold for i in V + ['t'])                       # min battery at customers
        self.m.addConstrs(
            self.b[j] <= self.b[i] - (a.get((i,j),a[i,0])/EV_velocity)*(gamma+gamma_l*self.f.get((i,j),self.f[i,0])) + (1-self.x[i,j])
            for i in V for j in N + ['t'] if (i != j and (i!=0 and j!='t'))
            )  # battery depletion
        
        self.m.setObjective(
            quicksum(w_ev*a[i,j]*self.x[i,j]  for i in V for j in V if i != j)   # base distance cost
            + (theta-self.dual_values_subsidy)* quicksum(260*EV_cost*(a[i,j]/EV_velocity)*(gamma+gamma_l*(self.f[i,j])) for i in V for j in V if i != j)
            - quicksum(self.dual_values_delta[i]*self.y[i] for i in N)
            - self.dual_values_vehicle
            - quicksum(self.dual_values_IR[i]*self.y[i]* (a[i,0]*GV_cost*q[i]+a[i,0]*GV_cost) for i in N),
            GRB.MINIMIZE
        )
        

        # Allow Gurobi to search for multiple solutions
        self.m.setParam("OutputFlag", 1)

        self.m.optimize()
        if self.m.Status == GRB.INFEASIBLE:
            print("Model is infeasible; computing IIS...")
            self.m.computeIIS()
            self.m.write("model.ilp")

        results = []

        if self.m.SolCount > 0:
            for k in range(self.m.SolCount):
                self.m.setParam(GRB.Param.SolutionNumber, k)
                obj_val = self.m.PoolObjVal
                if obj_val < -tol and abs(obj_val) > 0.001:
                    # Extract tour
                    tour = [0]
                    current = 0
                    while True:
                        next_nodes = [j for j in V if j != current and self.x[current, j].Xn > 0.5]
                        if not next_nodes:
                            break
                        nxt = next_nodes[0]
                        tour.append(nxt)
                        if nxt == 0:
                            break
                        current = nxt

                    results.append(tour)

        return results

    def rg_pctsp(self):
        """
        Prize-Collecting TSP with load-dependent travel costs.
        Flow-based formulation (no big-M load variables).
        Collects all negative-valued solutions.
        """

        # Map prizes to nodes
        prizes = {i: self.p_result.get(f"p_{i}", 0.0) for i in N}
        prizes[0] = 0.0  # depot has no prize

        self.m = self.pctsp()


        # Objective = base distance cost + load*distance cost – collected prizes
        self.m.setObjective(
            quicksum(a[0, j] * self.x[0, j] * GV_cost for j in N)   # base distance cost
            + quicksum(a[i, j] * self.f[i, j] * GV_cost for i in V for j in V if i != j) # load * distance cost
            - quicksum(prizes[i] * self.y[i] for i in V),                                # collected prizes
            GRB.MINIMIZE
        )

        # Allow Gurobi to search for multiple solutions
        self.m.setParam("OutputFlag", 1)

        self.m.optimize()

        results = []

        if self.m.SolCount > 0:
            for k in range(self.m.SolCount):
                self.m.setParam(GRB.Param.SolutionNumber, k)
                obj_val = self.m.PoolObjVal
                if obj_val < -tol and abs(obj_val) > 0.001:
                    # Extract tour
                    tour = [0]
                    current = 0
                    while True:
                        next_nodes = [j for j in V if j != current and self.x[current, j].Xn > 0.5]
                        if not next_nodes:
                            break
                        nxt = next_nodes[0]
                        tour.append(nxt)
                        if nxt == 0:
                            break
                        current = nxt

                    travel_cost = gv_tsp_cost(tour)
                    collected_prizes = sum(prizes[i] for i in tour)

                    results.append((tour, obj_val, travel_cost, collected_prizes))

        return results

@dataclass
class CGResult:
    y_r_result: dict
    not_fractional: bool
    model: any
    objval: float
    status: int
    CG_iteration: int
    RG_iteration: int
    RG_time: float
    CG_time: float
    CG_DP_time: float
    RG_DP_time: float
    LP_time: float
    tsp_memo: dict
    feasibility_memo: dict
    global_tsp_memo: dict
    num_lp: int
    new_constraints: set

def unpack_result(res: CGResult):
    return (
        res.y_r_result, res.not_fractional, res.model, res.objval, res.status,
        res.CG_iteration, res.RG_iteration, res.RG_time, res.CG_time, res.CG_DP_time,
        res.RG_DP_time, res.LP_time, res.tsp_memo, res.feasibility_memo,
        res.global_tsp_memo, res.num_lp, res.new_constraints
    )

