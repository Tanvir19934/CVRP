from config_new import q, a, EV_velocity, gamma, gamma_l, EV_cost, GV_cost, w_dv, w_ev, theta
from gurobipy import Model, GRB, quicksum
import itertools
from itertools import permutations
from openpyxl import load_workbook
import pandas as pd
import os
import importlib
import re


def refresh_config():
    import config_new
    importlib.reload(config_new)  # Reload the module to update V, q, a, Q_EV
    globals().update({k: getattr(config_new, k) for k in dir(config_new) if not k.startswith("__")})
    pass

def ev_travel_cost(route):
    refresh_config()
    b = 1
    l = 0
    for i in range(len(route)-1):
        l+=q[route[i]]
        b = b - (a[route[i],route[i+1]]/EV_velocity)*(gamma+gamma_l*l) 
    cost = 260*EV_cost*(1-b)
    return cost

def gv_tsp_cost(route):
    refresh_config()
    cost_GV = a[(route[0],route[1])]*GV_cost
    l = 0
    try:
        route = eval(route)
    except:
        for i in range(1,len(route)-1):
            l+=q[route[i]]
            cost_GV += a[(route[i],route[i+1])]*GV_cost*l
        return cost_GV
    
def check_values(d):
    for key, value in d.items():
        if value != 0 and value!=1:
            return False
    return True

def partial_path(label,current_node):
    node = current_node
    path = []
    while label:
        path.append(label.node)
        label = label.parent
    if node in path[1:]:
        return True
    else: return False

def reconstruct_path(label):
    path = []
    while label:
        path.append(label.node)
        label = label.parent
    path = path[::-1]
    if path[0]=='s':
        path[0] = 0
    if path[-1]=='t':
        path[-1]= 0
    return path

def construct_tour(edges):
    # Build a dictionary for quick lookup of the next node
    next_node_map = {i: j for i, j in edges}
    
    # Start from any node (choosing the first node from the first edge)
    start_node = edges[0][0]
    route = [start_node]
    
    current_node = start_node
    while True:
        next_node = next_node_map.get(current_node)
        if next_node is None or next_node == start_node:  # End of cycle
            break
        route.append(next_node)
        current_node = next_node

    route.append(start_node)  # Complete the cycle
    return route

def print_solution(final_model) -> None:
    refresh_config()
    model_vars = {var.VarName: var.X for var in final_model.getVars()}
    var_names = list(model_vars.keys())

    solution_routes = []
    payments = {}

    for var in var_names:
        value = round(model_vars[var], 2) # Get the variable's value

        if var.startswith("p_"):  # Payment variables
            payments[int(var.split("_")[-1])] = value  # Convert index to int
        elif value != 0 and var.startswith("y_r_["):  # Route variables
            print(f"{var}: {value}")
            solution_routes.append(list(eval(var[5:-1])))  # Extract tuple inside []

    # Print results
    c_r = {tuple(item):0 for item in solution_routes}
    for item in solution_routes:
        if len(item) > 3:
            c_r[tuple(item)] = ev_travel_cost(item)

    total_dv_miles_traveled = 0
    total_ev_miles_traveled = 0

    for route in solution_routes:
        if len(route) == 3:
            for i in range(0,len(route)-1):
                total_dv_miles_traveled += a[(route[i], route[i+1])]
        else:
            for i in range(0,len(route)-1):
                total_ev_miles_traveled += a[(route[i], route[i+1])]
    print(solution_routes)
    print(f"Total miles cost (ev+dv): {total_dv_miles_traveled*w_dv + total_ev_miles_traveled*w_ev}")
    print(f"Total EV miles traveled: {total_ev_miles_traveled}")
    print(f"Total DV miles traveled: {total_dv_miles_traveled}")
    print(f"Total mniles traveled: {total_dv_miles_traveled + total_ev_miles_traveled}")
    print(f"Objective value: {final_model.getObjective().getValue()}")
    print(f"Objective value (manual): {total_dv_miles_traveled*w_dv + total_ev_miles_traveled*w_ev+theta*(sum(c_r.values())-sum(payments.values()))}")
    print("Total payment received:", sum(payments.values()))
    print("Total coalition cost:", sum(c_r.values()))
    print("Total Subsidy:", sum(c_r.values())-sum(payments.values()))
    print(f"Payments: {payments}")

    return final_model.getObjective().getValue(), total_dv_miles_traveled*w_dv + total_ev_miles_traveled*w_ev \
        , total_ev_miles_traveled, sum(payments.values()), sum(c_r.values())-sum(payments.values()), payments, solution_routes

def generate_all_possible_routes(N):
    
    all_routes = []

    def generate_k_degree_coalition(N, k):
        # Generate all combinations of k nodes
        combinations = list(itertools.combinations(N, k))
        # Generate all possible routes starting and ending at depot 0
        degree_k_coalition = [tuple([0] + list(comb) + [0]) for comb in combinations]
        return degree_k_coalition
    

    for item in N:
        all_routes.extend(generate_k_degree_coalition(N, item))
    
    return all_routes
def tsp_tour(route):
    
    if len(route) == 3 and route[0]==0:
        return a[(0,route[1])]* GV_cost + a[(route[1],0)] * q[route[1]] * GV_cost, route

    intermediate_nodes = route[1:-1]

    all_routes = [[0] + list(p) + [0] for p in permutations(intermediate_nodes)]

    routes_list = [tuple(all_routes) for all_routes in all_routes]

    route_cost = {}

    for item in routes_list:
        route_cost[tuple(item)] = gv_tsp_cost(item)

    model = Model("TSP")
    x = model.addVars(routes_list, vtype=GRB.BINARY, name="x")

    model.addConstr(quicksum(x[i] for i in routes_list) == 1)

    model.setObjective(quicksum(route_cost[i] * x[i] for i in routes_list), GRB.MINIMIZE)


    model.Params.OutputFlag = 0
    model.optimize()

    # Extract the solution
    if model.status == GRB.OPTIMAL:
        for i in routes_list:
            if x[i].X > 0.5:
                tour = i
                break
        return model.getObjective().getValue(), tour
    


def save_to_excel(file_name, sheet_name, data):
    try:
        # Check if the file already exists
        if os.path.exists(file_name):
            # Load the existing workbook
            book = load_workbook(file_name)
            # Check if the sheet already exists
            if sheet_name in book.sheetnames:
                # Read the existing data in the sheet
                existing_data = pd.read_excel(file_name, sheet_name=sheet_name)
                # Append the new data to the existing data
                updated_data = pd.concat([existing_data, data], ignore_index=True)
            else:
                # If the sheet doesn't exist, use the new data
                updated_data = data

            # Write the updated data to the sheet
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                updated_data.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # If the file doesn't exist, create a new one with the data
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Data saved to {file_name} in sheet '{sheet_name}'.")
    except Exception as e:
        print(f"Error: {e}")


def update_config(node_value):
    # Path to the config_new.py file
    config_file = "New_codes/config_new.py"

    # Step 1: Update the NODES value in the config_new.py file
    with open(config_file, "r") as f:
        lines = f.readlines()

    with open(config_file, "w") as f:
        for line in lines:
            if line.startswith("NODES ="):
                f.write(f"NODES = {node_value}\n")  # Update NODES value
            else:
                f.write(line)

    # Step 3: Re-execute the rest of the code in the config_new module
    # This will execute the entire code from config_new.py again
    os.system('python3 New_codes/config_new.py')  # Executes the script after updating NODES


def create_excel_for_log_file(log_file):

    # Regular expression pattern to extract timestamp, description, and time value
    log_pattern = r"^(.*?) - (.*?): (\d+\.\d+)"

    # Lists to store parsed data
    timestamps = []
    descriptions = []
    time_values = []

    # Read the log file and extract data
    with open(log_file, "r") as file:
        for line in file:
            match = re.search(log_pattern, line)
            if match:
                timestamps.append(match.group(1))      # Timestamp
                descriptions.append(match.group(2))    # Description
                time_values.append(float(match.group(3)))  # Time value

    # Create a DataFrame
    df = pd.DataFrame({
        "Timestamp": timestamps,
        "Description": descriptions,
        "Time Taken": time_values
    })

    # Save to an Excel file
    log_file = log_file.split(".")[0]
    excel_filename = f"{log_file}.xlsx"
    df.to_excel(excel_filename, index=False)

    print(f"Data successfully saved to {excel_filename}")